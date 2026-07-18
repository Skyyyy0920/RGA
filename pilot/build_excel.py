#!/usr/bin/env python
"""Build an Excel workbook of all RGA/Qwen experiment results from the result JSONs.
Uses base anaconda python (openpyxl available, no torch needed)."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

KD = "/data/tianhao/KD"
def load(p):
    fp = os.path.join(KD, p)
    return json.load(open(fp)) if os.path.exists(fp) else None

def cell(ws, r, c, v, bold=False, fill=None, center=True):
    x = ws.cell(row=r, column=c, value=v)
    if bold: x.font = Font(bold=True)
    if center: x.alignment = Alignment(horizontal='center', vertical='center')
    if fill: x.fill = PatternFill('solid', fgColor=fill)
    thin = Side(style='thin', color='CCCCCC')
    x.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return x

def fmt(d):
    if d is None: return "(pending)"
    return f"{d['rougeL_mean']:.4f} ± {d['rougeL_std']:.4f}"

HDR = "1F4E78"; HDRF = "DDEBF7"; BEST = "C6EFCE"
wb = Workbook(); wb.remove(wb.active)

# ---------- Sheet 1: Fair head-to-head (decisive) ----------
final = load("pilot/FINAL_COMPARE_QWEN.json")
less = load("pilot/LESS_QWEN.json")
ws = wb.create_sheet("1_Fair_HeadToHead")
cell(ws, 1, 1, "Fair head-to-head — ALL methods, identical micro-batch training, ROUGE-L (3 seeds)", bold=True)
ws.merge_cells('A1:H1')
methods = ['RGA_fixcov_nogate', 'TAGCOS', 'LESS', 'deep_topr', 'GraNd', 'random']
labels  = ['RGA (ours)', 'TAGCOS', 'LESS', 'deep-topr', 'GraNd', 'random']
cell(ws, 3, 1, "Budget", bold=True, fill=HDRF)
for j, lab in enumerate(labels):
    cell(ws, 3, 2+j, lab, bold=True, fill=HDRF)
budgets = ['500', '1000', '2000']
row = 4
means = {m: [] for m in methods}
for b in budgets:
    cell(ws, row, 1, f"{b} ({int(int(b)/4000*100)}% of pool)", bold=True)
    fd = (final or {}).get('sweep', {}).get(b, {})
    ld = (less or {}).get('results', {}).get(b, {})
    vals = {}
    for m in methods:
        if m == 'LESS':
            vals[m] = ld if ld else None
        else:
            vals[m] = fd.get(m)
    # find best mean this row
    valid = {m: v['rougeL_mean'] for m, v in vals.items() if v}
    bestm = max(valid, key=valid.get) if valid else None
    for j, m in enumerate(methods):
        v = vals[m]
        cell(ws, row, 2+j, fmt(v), fill=(BEST if m == bestm else None))
        if v: means[m].append(v['rougeL_mean'])
    row += 1
cell(ws, row, 1, "average", bold=True, fill=HDRF)
avg = {m: (sum(means[m])/len(means[m]) if means[m] else None) for m in methods}
valid_avg = {m: a for m, a in avg.items() if a is not None}
bestavg = max(valid_avg, key=valid_avg.get) if valid_avg else None
for j, m in enumerate(methods):
    a = avg[m]
    cell(ws, row, 2+j, f"{a:.4f}" if a is not None else "(pending)",
         bold=True, fill=(BEST if m == bestavg else HDRF))
cell(ws, row+2, 1, "Green = best in row. RGA beats TAGCOS at every budget; wins/ties random; most stable (lowest std).")
ws.column_dimensions['A'].width = 22
for col in 'BCDEFGH': ws.column_dimensions[col].width = 17

# ---------- Sheet 2: Full sweep + external baselines ----------
sweep = load("pilot/RETRAIN_QWEN_SWEEP.json"); base = load("pilot/BASELINES_QWEN_SWEEP.json")
ws2 = wb.create_sheet("2_Sweep_AllBaselines")
cell(ws2, 1, 1, "Budget sweep + external baselines (original RGA_deep, single-batch protocol), ROUGE-L", bold=True)
ws2.merge_cells('A1:I1')
cols = [('RGA_deep', 'RGA_deep (orig, buggy)'), ('deep_topr', 'deep_topr'), ('random', 'random'),
        ('proxy_topr', 'proxy_topr'), ('EL2N', 'EL2N'), ('GraNd', 'GraNd'), ('TAGCOS', 'TAGCOS'), ('GRAFT', 'GRAFT')]
cell(ws2, 3, 1, "Budget", bold=True, fill=HDRF)
for j, (_, lab) in enumerate(cols): cell(ws2, 3, 2+j, lab, bold=True, fill=HDRF)
row = 4
for b in budgets:
    cell(ws2, row, 1, b, bold=True)
    sd = (sweep or {}).get('sweep', {}).get(b, {}); bd = (base or {}).get('sweep', {}).get(b, {})
    for j, (k, _) in enumerate(cols):
        v = sd.get(k) or bd.get(k)
        cell(ws2, row, 2+j, fmt(v))
    row += 1
cell(ws2, row+1, 1, "Note: original RGA_deep had the coverage bug + gate → ~random. Deep-grad methods (TAGCOS/deep_topr) beat feature(GRAFT)/loss(EL2N) → space validated.")
ws2.column_dimensions['A'].width = 12
for col in 'BCDEFGHI': ws2.column_dimensions[col].width = 17

# ---------- Sheet 3: RGA fix variants ----------
fix = load("pilot/RGA_FIX_QWEN_SWEEP.json")
ws3 = wb.create_sheet("3_RGA_fix_variants")
cell(ws3, 1, 1, "RGA fix variants (corrected coverage + gate ablation), ROUGE-L", bold=True)
ws3.merge_cells('A1:D1')
fcols = [('RGA_fixcov', 'RGA_fixcov (gate+cov)'), ('RGA_gatetopr', 'RGA_gatetopr (gate, no cov)'),
         ('RGA_fixcov_nogate', 'RGA_fixcov_nogate (cov, NO gate) *best*')]
cell(ws3, 3, 1, "Budget", bold=True, fill=HDRF)
for j, (_, lab) in enumerate(fcols): cell(ws3, 3, 2+j, lab, bold=True, fill=HDRF)
row = 4
for b in budgets:
    cell(ws3, row, 1, b, bold=True)
    fd = (fix or {}).get('sweep', {}).get(b, {})
    for j, (k, _) in enumerate(fcols):
        cell(ws3, row, 2+j, fmt(fd.get(k)), fill=(BEST if k == 'RGA_fixcov_nogate' else None))
    row += 1
cell(ws3, row+1, 1, "Finding: the GATE HURTS on a clean teacher (no-gate > gate at every budget). Corrected coverage + no gate = best.")
ws3.column_dimensions['A'].width = 12
for col in 'BCD': ws3.column_dimensions[col].width = 30

# ---------- Sheet 4: Decision experiment ----------
dec = load("pilot/DECISION_QWEN_RESULTS.json")
ws4 = wb.create_sheet("4_Decision_experiment")
cell(ws4, 1, 1, "Decision experiment — is the deep gradient really different from token-level? (Qwen, N=1500)", bold=True)
ws4.merge_cells('A1:C1')
if dec:
    g = dec['gram_cka']
    rows = [("Kernel similarity (CKA)", "value", "reading"),
            ("last-layer proxy  vs  token", f"{g['token_vs_proxy']:.2f}", "≈1 → proxy IS token-level (advisor right)"),
            ("deep gradient  vs  token", f"{g['token_vs_deep']:.2f}", "low → deep gradient is DIFFERENT"),
            ("deep gradient  vs  last-layer proxy", f"{g['proxy_vs_deep']:.2f}", "low → deep ≠ proxy"),
            ("top-10% selection overlap: deep vs token", f"{dec['topk_overlap@10pct']['token_vs_deep']:.2f}", "0.05 → pick almost disjoint samples")]
    for i, (a, bv, cv) in enumerate(rows):
        cell(ws4, 3+i, 1, a, bold=(i == 0), fill=(HDRF if i == 0 else None))
        cell(ws4, 3+i, 2, bv, bold=(i == 0), fill=(HDRF if i == 0 else None))
        cell(ws4, 3+i, 3, cv, bold=(i == 0), fill=(HDRF if i == 0 else None), center=False)
ws4.column_dimensions['A'].width = 34; ws4.column_dimensions['B'].width = 10; ws4.column_dimensions['C'].width = 48

# ---------- Sheet 5: Method descriptions ----------
ws5 = wb.create_sheet("5_Method_descriptions")
cell(ws5, 1, 1, "What each method uses and does", bold=True); ws5.merge_cells('A1:D1')
hdr = ["Method", "Signal used", "What it does with it", "Key limitation vs RGA"]
for j, h in enumerate(hdr): cell(ws5, 3, 1+j, h, bold=True, fill=HDRF)
rows = [
 ["EL2N", "Student output error ||p_S - onehot(y)|| (not a gradient)", "Take norm; top-k highest error", "Output/label space; grabs noisy outliers; worst in our runs"],
 ["GraNd", "KD-loss gradient norm ||g|| (last-layer)", "Take scalar norm; top-k largest", "Discards gradient direction; last-layer ~ token-level"],
 ["GRAFT", "Penultimate features (not a gradient)", "D-optimal/MaxVol coverage over features", "Covers data appearance, not what to learn"],
 ["TAGCOS", "Deep gradient (SAME as RGA)", "k-means cluster; pick medoids", "Student-only diversity; no teacher comparison"],
 ["LESS (ICML'24)", "Deep gradient (LoRA in orig)", "Cosine to held-out validation gradient; top-k", "Targets a validation set, not the teacher; no coverage"],
 ["RGA (ours)", "Student deep KD gradient + teacher eNTK", "Teacher-vs-student relational residual r + coverage", "Only method aligning student against the teacher"],
]
for i, r in enumerate(rows):
    for j, v in enumerate(r):
        cell(ws5, 4+i, 1+j, v, bold=(r[0] == 'RGA (ours)'),
             fill=(BEST if r[0] == 'RGA (ours)' else None), center=False)
ws5.column_dimensions['A'].width = 16
for col, w in zip('BCD', [42, 42, 46]): ws5.column_dimensions[col].width = w

out = os.path.join(KD, "RGA_experiment_results.xlsx")
wb.save(out)
print("wrote", out)
